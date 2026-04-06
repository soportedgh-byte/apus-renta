"""
CecilIA v2 — Exportador de reportes analiticos
Contraloria General de la Republica de Colombia

Archivo: app/analytics/exportador.py
Proposito: Exportacion a Excel (openpyxl), DOCX informe ejecutivo, CSV.
Sprint: 10
Autor: Equipo Tecnico CecilIA — CD-TIC-CGR
Fecha: Abril 2026
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger("cecilia.analytics.exportador")


# ── Exportacion a Excel ─────────────────────────────────────────────────────

def exportar_excel(
    datos_uso: dict[str, Any],
    datos_auditoria: dict[str, Any],
    datos_calidad: dict[str, Any],
    datos_capacitacion: dict[str, Any],
) -> bytes:
    """Genera un archivo Excel con multiples hojas de metricas.

    Returns:
        Bytes del archivo .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Estilos
    encabezado_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    encabezado_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    oro_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _escribir_hoja(ws, titulo: str, filas: list[list[Any]], encabezados: list[str]):
        ws.title = titulo
        # Encabezados
        for j, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=j, value=enc)
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal="center")
            celda.border = borde
        # Datos
        for i, fila in enumerate(filas, 2):
            for j, valor in enumerate(fila, 1):
                celda = ws.cell(row=i, column=j, value=valor)
                celda.border = borde
        # Ajustar anchos
        for col in ws.columns:
            max_len = max(len(str(celda.value or "")) for celda in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Hoja 1: Uso
    ws_uso = wb.active
    _escribir_hoja(ws_uso, "Uso del Sistema", [
        ["Total consultas", datos_uso.get("total_consultas", 0)],
        ["Usuarios activos", datos_uso.get("total_usuarios_activos", 0)],
        ["Consultas hoy", datos_uso.get("consultas_hoy", 0)],
        ["Consultas semana", datos_uso.get("consultas_semana", 0)],
        ["Consultas mes", datos_uso.get("consultas_mes", 0)],
        ["Consultas DES", datos_uso.get("consultas_por_direccion", {}).get("DES", 0)],
        ["Consultas DVF", datos_uso.get("consultas_por_direccion", {}).get("DVF", 0)],
    ], ["Metrica", "Valor"])

    # Hoja 2: Auditoria
    ws_aud = wb.create_sheet()
    filas_aud = [
        ["Total auditorias", datos_auditoria.get("total_auditorias", 0)],
        ["Total hallazgos", datos_auditoria.get("total_hallazgos", 0)],
        ["Hallazgos IA", datos_auditoria.get("hallazgos_generados_ia", 0)],
        ["Hallazgos validados", datos_auditoria.get("hallazgos_validados_humano", 0)],
        ["Tasa aprobacion (%)", datos_auditoria.get("tasa_hallazgos_aprobados", 0)],
        ["Total formatos", datos_auditoria.get("total_formatos", 0)],
        ["Formatos IA", datos_auditoria.get("formatos_generados_ia", 0)],
        ["Cuantia presunto dano ($)", datos_auditoria.get("cuantia_total_presunto_dano", 0)],
    ]
    # Connotaciones
    for con, val in datos_auditoria.get("hallazgos_por_connotacion", {}).items():
        filas_aud.append([f"Hallazgos {con}", val])
    _escribir_hoja(ws_aud, "Auditorias", filas_aud, ["Metrica", "Valor"])

    # Hoja 3: Calidad
    ws_cal = wb.create_sheet()
    _escribir_hoja(ws_cal, "Calidad", [
        ["Feedback positivo", datos_calidad.get("feedback_positivo", 0)],
        ["Feedback negativo", datos_calidad.get("feedback_negativo", 0)],
        ["Tasa satisfaccion (%)", datos_calidad.get("tasa_satisfaccion", 0)],
        ["Cobertura fuentes (%)", datos_calidad.get("cobertura_fuentes", 0)],
        ["Total respuestas", datos_calidad.get("total_respuestas", 0)],
        ["Hallazgos aprobados", datos_calidad.get("hallazgos_aprobados", 0)],
        ["Tasa aprobacion hallazgos (%)", datos_calidad.get("tasa_hallazgos_aprobados", 0)],
    ], ["Metrica", "Valor"])

    # Hoja 4: Capacitacion
    ws_cap = wb.create_sheet()
    _escribir_hoja(ws_cap, "Capacitacion", [
        ["Funcionarios capacitados", datos_capacitacion.get("total_funcionarios_capacitados", 0)],
        ["Lecciones completadas", datos_capacitacion.get("lecciones_completadas", 0)],
        ["Quizzes realizados", datos_capacitacion.get("total_quizzes", 0)],
        ["Quizzes aprobados", datos_capacitacion.get("quizzes_aprobados", 0)],
        ["Tasa aprobacion (%)", datos_capacitacion.get("tasa_aprobacion_quizzes", 0)],
        ["Promedio puntaje", datos_capacitacion.get("promedio_puntaje", 0)],
        ["Rutas activas", datos_capacitacion.get("rutas_activas", 0)],
    ], ["Metrica", "Valor"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── Exportacion a CSV ────────────────────────────────────────────────────────

def exportar_csv(datos: dict[str, Any], seccion: str = "uso") -> bytes:
    """Genera un archivo CSV con las metricas de una seccion.

    Args:
        datos: Dict de metricas.
        seccion: Nombre de la seccion para el header.

    Returns:
        Bytes del archivo CSV.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Metrica", "Valor", "Seccion"])

    for clave, valor in datos.items():
        if isinstance(valor, dict):
            for sub_clave, sub_valor in valor.items():
                writer.writerow([f"{clave}.{sub_clave}", sub_valor, seccion])
        elif isinstance(valor, list):
            writer.writerow([clave, len(valor), seccion])
        else:
            writer.writerow([clave, valor, seccion])

    return buffer.getvalue().encode("utf-8-sig")


# ── Informe ejecutivo DOCX ──────────────────────────────────────────────────

def generar_informe_ejecutivo(
    datos_uso: dict[str, Any],
    datos_auditoria: dict[str, Any],
    datos_calidad: dict[str, Any],
    datos_capacitacion: dict[str, Any],
    periodo: str = "",
) -> bytes:
    """Genera un informe ejecutivo mensual en DOCX con branding CGR.

    Returns:
        Bytes del archivo DOCX.
    """
    from app.formatos.formato_base import FormatoBaseCGR

    class InformeEjecutivo(FormatoBaseCGR):
        def __init__(self, uso, auditoria, calidad, capacitacion, periodo):
            super().__init__(
                numero_formato=98,
                nombre_formato="Informe Ejecutivo Mensual — CecilIA v2",
                datos={
                    "nombre_entidad": "Contraloria General de la Republica",
                    "vigencia": datetime.now().strftime("%Y"),
                    "tipo_auditoria": "Informe de Gestion IA",
                },
            )
            self.uso = uso
            self.auditoria = auditoria
            self.calidad = calidad
            self.capacitacion = capacitacion
            self.periodo_str = periodo or datetime.now().strftime("%B %Y")

        def construir(self) -> None:
            # 1. Resumen ejecutivo
            self.agregar_titulo_seccion("1. Resumen Ejecutivo")
            self.agregar_parrafo_justificado(
                f"El presente informe consolida las metricas de operacion del sistema "
                f"CecilIA v2 durante el periodo {self.periodo_str}. El sistema atendio "
                f"{self.uso.get('total_consultas', 0):,} consultas de "
                f"{self.uso.get('total_usuarios_activos', 0)} usuarios activos."
            )

            # KPIs principales
            self.agregar_titulo_seccion("2. Indicadores Clave (KPIs)")
            self.crear_tabla(
                ["Indicador", "Valor", "Observacion"],
                [
                    ["Consultas totales", f"{self.uso.get('total_consultas', 0):,}", "Mensajes de usuario"],
                    ["Usuarios activos", str(self.uso.get("total_usuarios_activos", 0)), "Usuarios unicos"],
                    ["Satisfaccion", f"{self.calidad.get('tasa_satisfaccion', 0):.0f}%", "Basado en feedback"],
                    ["Hallazgos generados", str(self.auditoria.get("total_hallazgos", 0)), "Con asistencia IA"],
                    ["Formatos producidos", str(self.auditoria.get("total_formatos", 0)), "Formatos CGR"],
                    ["Funcionarios capacitados", str(self.capacitacion.get("total_funcionarios_capacitados", 0)), "Modulo tutor"],
                ],
                anchos=[5, 3, 7],
            )

            # 3. Uso por direccion
            self.agregar_titulo_seccion("3. Comparativo DES vs DVF")
            des = self.uso.get("consultas_por_direccion", {}).get("DES", 0)
            dvf = self.uso.get("consultas_por_direccion", {}).get("DVF", 0)
            self.crear_tabla(
                ["Direccion", "Consultas", "Participacion"],
                [
                    ["DES", str(des), f"{des / max(des + dvf, 1) * 100:.0f}%"],
                    ["DVF", str(dvf), f"{dvf / max(des + dvf, 1) * 100:.0f}%"],
                ],
                anchos=[5, 4, 4],
            )

            # 4. Auditoria
            self.agregar_titulo_seccion("4. Metricas de Auditoria")
            connotaciones = self.auditoria.get("hallazgos_por_connotacion", {})
            self.crear_tabla(
                ["Connotacion", "Hallazgos"],
                [[k.capitalize(), str(v)] for k, v in connotaciones.items()],
                anchos=[6, 3],
            )
            cuantia = self.auditoria.get("cuantia_total_presunto_dano", 0)
            if cuantia:
                self.agregar_parrafo(
                    f"Cuantia total de presunto dano patrimonial identificado: "
                    f"${cuantia:,.0f} millones.",
                    negrita=True,
                )

            # 5. Calidad
            self.agregar_titulo_seccion("5. Calidad de Respuestas")
            self.crear_tabla_clave_valor([
                ("Tasa de satisfaccion", f"{self.calidad.get('tasa_satisfaccion', 0):.1f}%"),
                ("Cobertura de fuentes", f"{self.calidad.get('cobertura_fuentes', 0):.1f}%"),
                ("Feedback positivo", str(self.calidad.get("feedback_positivo", 0))),
                ("Feedback negativo", str(self.calidad.get("feedback_negativo", 0))),
                ("Hallazgos aprobados", f"{self.calidad.get('tasa_hallazgos_aprobados', 0):.1f}%"),
            ])

            # 6. Capacitacion
            self.agregar_titulo_seccion("6. Modulo de Capacitacion")
            self.crear_tabla_clave_valor([
                ("Funcionarios capacitados", str(self.capacitacion.get("total_funcionarios_capacitados", 0))),
                ("Lecciones completadas", str(self.capacitacion.get("lecciones_completadas", 0))),
                ("Quizzes aprobados", str(self.capacitacion.get("quizzes_aprobados", 0))),
                ("Tasa aprobacion", f"{self.capacitacion.get('tasa_aprobacion_quizzes', 0):.1f}%"),
                ("Promedio puntaje", f"{self.capacitacion.get('promedio_puntaje', 0):.1f}/100"),
            ])

            # 7. Recomendaciones
            self.agregar_titulo_seccion("7. Recomendaciones")
            recomendaciones = []
            if self.calidad.get("tasa_satisfaccion", 0) < 80:
                recomendaciones.append(
                    "Revisar las respuestas con feedback negativo y ajustar los prompts del sistema."
                )
            if self.calidad.get("cobertura_fuentes", 0) < 70:
                recomendaciones.append(
                    "Ampliar la base documental RAG para mejorar la cobertura de citaciones."
                )
            if self.capacitacion.get("total_funcionarios_capacitados", 0) < 10:
                recomendaciones.append(
                    "Promover el uso del modulo de capacitacion entre los auditores."
                )
            if not recomendaciones:
                recomendaciones.append(
                    "El sistema opera dentro de los parametros esperados. "
                    "Continuar el monitoreo mensual."
                )
            for i, rec in enumerate(recomendaciones, 1):
                self.agregar_parrafo(f"{i}. {rec}")

    informe = InformeEjecutivo(
        datos_uso, datos_auditoria, datos_calidad, datos_capacitacion, periodo,
    )
    return informe.generar_bytes()


# ── Informe Circular 023 DOCX ───────────────────────────────────────────────

def generar_reporte_circular_023_docx(datos: dict[str, Any]) -> bytes:
    """Genera el reporte trimestral Circular 023 en formato DOCX."""
    from app.formatos.formato_base import FormatoBaseCGR

    class ReporteC023(FormatoBaseCGR):
        def __init__(self, datos):
            super().__init__(
                numero_formato=97,
                nombre_formato="Reporte Circular 023 — Uso de IA",
                datos={
                    "nombre_entidad": "Contraloria General de la Republica",
                    "vigencia": datos.get("periodo_trimestre", ""),
                    "tipo_auditoria": "Informe de Cumplimiento Circular 023",
                },
            )
            self.datos_reporte = datos

        def construir(self) -> None:
            d = self.datos_reporte

            self.agregar_titulo_seccion("1. Informacion General")
            self.crear_tabla_clave_valor([
                ("Periodo", d.get("periodo_trimestre", "")),
                ("Fecha inicio", str(d.get("periodo_inicio", ""))[:10]),
                ("Fecha fin", str(d.get("periodo_fin", ""))[:10]),
            ])

            self.agregar_titulo_seccion("2. Estadisticas de Uso")
            self.crear_tabla(
                ["Metrica", "Valor"],
                [
                    ["Total consultas", f"{d.get('total_consultas', 0):,}"],
                    ["Total conversaciones", f"{d.get('total_conversaciones', 0):,}"],
                    ["Promedio mensajes/conversacion", f"{d.get('promedio_mensajes_por_conversacion', 0):.1f}"],
                    ["Usuarios activos DES", str(d.get("usuarios_activos_des", 0))],
                    ["Usuarios activos DVF", str(d.get("usuarios_activos_dvf", 0))],
                    ["Documentos procesados RAG", str(d.get("documentos_procesados_rag", 0))],
                    ["Formatos generados con IA", str(d.get("formatos_generados_ia", 0))],
                    ["Hallazgos asistidos por IA", str(d.get("hallazgos_asistidos_ia", 0))],
                ],
                anchos=[8, 5],
            )

            self.agregar_titulo_seccion("3. Calidad y Retroalimentacion")
            self.crear_tabla(
                ["Indicador", "Valor"],
                [
                    ["Feedback positivo", str(d.get("feedback_positivo", 0))],
                    ["Feedback negativo", str(d.get("feedback_negativo", 0))],
                    ["Tasa satisfaccion", f"{d.get('tasa_satisfaccion', 0):.1f}%"],
                ],
                anchos=[8, 5],
            )

            self.agregar_titulo_seccion("4. Cumplimiento Principios Circular 023")
            principios = d.get("principios_implementados", [])
            self.crear_tabla(
                ["Principio", "Estado"],
                [[p, "Implementado"] for p in principios],
                anchos=[8, 5],
            )

            self.agregar_parrafo(d.get("disclaimers_incluidos", ""))
            self.agregar_parrafo("")
            self.agregar_parrafo(d.get("nota", ""), negrita=True)

    reporte = ReporteC023(datos)
    return reporte.generar_bytes()
